import os
import random
from openpyxl import load_workbook

excel_file = r"C:\Users\KarinaKresnadi\Documents\General Document Mapping TEST.xlsx"  # Update this path
directory = r"M:\TMR\Sample Data\DMS\Sample_Data_Extract\PDF"  # Update this path

start_row = 3
n = 250

wb = load_workbook(excel_file)

sheet1 = wb['Data']  # Target 
sheet2 = wb['Sheet2']  # Source 


# Generate files and store their paths
file_paths = []
for i in range(1, n):
    file_path = os.path.normpath(f"{directory}\DMS Sample - 10 - TST - test({i+1}).pdf")  # Ensure consistent separators
    file_paths.append(file_path)

# os.makedirs(directory, exist_ok=True)

values = []

for row in range(2, 167):
    cell_value = sheet2.cell(row=row, column=3).value
    values.append(cell_value)

shuffled_values = values * (n // len(values)) + values[:n % len(values)]
random.shuffle(shuffled_values)

row = start_row
while sheet1.cell(row=row, column=6).value:
    row += 1
for row in range(3, n + 2): # do not ask me why its +2 this doesnt make sense to me
    random_value = shuffled_values.pop()  # Take a value from the shuffled list
    sheet1.cell(row=row, column=6, value=random_value)

print("Values have been randomly placed in Column Function:Sub Function.")



row = start_row
while sheet1.cell(row=row, column=3).value:
    row += 1
# start writing!!!!!!
for i, path in enumerate(file_paths, start=1):
    sheet1.cell(row=row, column=3, value=path)

    OTO_Title = f"DMS Sample - 10 - TST - test_{i+1}"
    sheet1.cell(row=row, column=4, value=OTO_Title)

    row += 1

wb.save(excel_file)
print("File paths recorded in Excel.")